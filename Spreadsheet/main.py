from os import close
import openpyxl
import json
import codecs

inv_file = openpyxl.load_workbook('./training-code/py-training/'
                                  'Spreadsheet/mobile_users.xlsx')
product_list = inv_file["Sheet1"]
excess_abonents = {}

try:
    for mob_abonent_row in range(2, product_list.max_row):
        #print(product_list.cell(mob_abonent_row, 7).value)
        abonent_name = product_list.cell(mob_abonent_row, 2).value
        excess_perc = product_list.cell(mob_abonent_row, 7).value
        if excess_perc > 0:
            excess_abonents[abonent_name] = '{} %'.format(excess_perc)
        else:
            None
    json_object = json.dumps(excess_abonents, indent=5, 
                             ensure_ascii=False)
    print(json_object)
    with codecs.open("sample.json", "w+", "utf-8") as outfile:
        outfile.write(json_object)
except TypeError:
    print("unsupported data in the sheet")

#print(excess_abonents)